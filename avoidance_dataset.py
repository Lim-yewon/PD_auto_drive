#!/usr/bin/env python3
"""
avoidance_dataset.py  ─ v2.0
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
자율주행 시뮬레이터 회피 기동 데이터셋 생성기
index.html GUI 와 물리 모델·상수 완전 동일

파라미터 = index.html GUI 슬라이더 3개
  핸들 조향각       steer_deg        −90 ~ +90 °  (핸들 SW 각도)
  회피 기동 시작점   straight_dist_m   5 ~ 120 m
  속도              speed_kmh        10 ~ 120 km/h

장애물 위치(JS DEFAULT_OBS 동일): x=0, z=−62 m

출력
  avoidance_dataset.csv   원시 시뮬레이션 결과
  avoidance_dataset.xlsx  엑셀 (원시 + 피벗 5장)
  avoidance_plots.png     시각화 그래프
  avoidance_report.txt    텍스트 요약 리포트

사용법
  python3 avoidance_dataset.py            # 전체 스윕
  python3 avoidance_dataset.py --quick    # 소규모 테스트
  python3 avoidance_dataset.py --workers 4

물리 단계 (index.html 완전 동일)
  STRAIGHT → EVADE(SR=15) → PD_RETURN → LANE_RETURN → COMPLETE
  낭떠러지: |x| > 7.5 m
  충돌 반경: 3.0 m (차량 1.5 + 장애물 1.5)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""

import math, csv, sys, os, time, multiprocessing
from typing import Dict, List, Tuple

import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from matplotlib.colors import LinearSegmentedColormap
import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter


# ══════════════════════════════════════════════════════════════
#  물리 상수  ── index.html 와 완전 동일
# ══════════════════════════════════════════════════════════════
WHEELBASE       = 2.5
MAX_STEER_RAD   = math.pi / 6          # 30 °
VEH_RADIUS      = 1.5
OBS_RADIUS      = 1.5
COLL_RADIUS     = VEH_RADIUS + OBS_RADIUS   # 3.0 m
CLIFF_X         = 7.5                  # 낭떠러지 경계 m
TARGET_THETA    = math.pi / 2          # 직진 헤딩
STEERING_RATIO  = 15                   # 핸들각 / 바퀴각  (JS 와 동일)
PHYS_DT         = 0.02                 # 50 Hz

EVADE_DIST      = 30.0                 # EVADE 구간 m
PD_KP           = 0.5
PD_KD           = 0.05
PD_VREF         = 60.0 / 3.6          # ≈ 16.67 m/s  (속도 적응 기준)

OBS_X           = 0.0                  # 장애물 고정 위치 (JS DEFAULT_OBS)
OBS_Z           = -62.0
MAX_TRAVEL      = 700.0                # 시뮬 최대 주행 거리 m


# ══════════════════════════════════════════════════════════════
#  파라미터 스윕 공간
# ══════════════════════════════════════════════════════════════
PARAM_FULL = {
    # JS 슬라이더 범위 그대로 — SW 각도 단위
    "steer_deg":        list(range(-90, 91, 5)),   # 37 개
    "straight_dist_m":  list(range(5,  121, 5)),   # 24 개  (5·10·…·120)
    "speed_kmh":        list(range(10, 130, 10)),  # 12 개  (10·20·…·120)
}

PARAM_QUICK = {
    "steer_deg":        list(range(-90, 91, 15)),  # 13 개
    "straight_dist_m":  [5, 15, 30, 45, 60, 90],
    "speed_kmh":        [20, 40, 60, 80, 100, 120],
}


# ══════════════════════════════════════════════════════════════
#  단일 시뮬레이션  ── index.html 물리 루프 완전 이식
# ══════════════════════════════════════════════════════════════
def simulate(args: Tuple) -> Dict:
    """
    (steer_deg_sw, straight_dist_m, speed_kmh) → 결과 딕셔너리
    index.html 의 4단계 물리 루프(STRAIGHT/EVADE/PD_RETURN/LANE_RETURN)를
    그대로 재현. 전역 함수 → multiprocessing.Pool pickle 가능.
    """
    steer_sw, straight_dist, speed_kmh = args

    v = speed_kmh / 3.6
    if v < 0.1:
        return _row(steer_sw, straight_dist, speed_kmh,
                    'STOPPED', 0, 999.0, 0.0, 0.0, -1, -1, 0.0)

    # ── 상태 변수 ──────────────────────────────────────────────
    x, z, theta = 0.0, 0.0, TARGET_THETA
    odist = 0.0
    phase = 'STRAIGHT'
    phase_start = 0.0
    prev_pd_err   = 0.0
    prev_lane_err = 0.0

    # EVADE 바퀴 조향각: SW → wheel (÷ SR) → rad  [JS update() 와 동일]
    raw_evade = -(steer_sw / STEERING_RATIO) * math.pi / 180.0
    delta_evade = max(-MAX_STEER_RAD, min(MAX_STEER_RAD, raw_evade))

    # 통계 추적
    min_obs_dist  = math.inf
    max_lat_dev   = 0.0
    lat_at_obs    = x
    passed_obs    = False
    pd_start_od   = -1.0
    lane_start_od = -1.0

    max_steps = int(MAX_TRAVEL / (v * PHYS_DT)) + 200

    for _ in range(max_steps):

        # ── 충돌 검사 (스텝 크기 ≤ 0.67 m ≪ 3.0 m → 포인트 샘플 ≈ CCD) ──
        d_obs = math.sqrt((x - OBS_X) ** 2 + (z - OBS_Z) ** 2)
        if d_obs < min_obs_dist:
            min_obs_dist = d_obs

        if not passed_obs and z <= OBS_Z:
            passed_obs  = True
            lat_at_obs  = x

        if d_obs < COLL_RADIUS:
            return _row(steer_sw, straight_dist, speed_kmh,
                        f'COLLISION_{phase}', 0,
                        min_obs_dist - COLL_RADIUS, max_lat_dev,
                        lat_at_obs if passed_obs else x,
                        pd_start_od, lane_start_od, odist)

        # 낭떠러지 검사  [JS: if(!sim.falling && Math.abs(egoModel.x) > CLIFF_X)]
        if abs(x) > CLIFF_X:
            return _row(steer_sw, straight_dist, speed_kmh,
                        f'CLIFF_{phase}', 0,
                        min_obs_dist - COLL_RADIUS, max_lat_dev,
                        lat_at_obs if passed_obs else x,
                        pd_start_od, lane_start_od, odist)

        # ── 단계별 조향 계산 ────────────────────────────────────

        # STRAIGHT: 직진  ─────────────────────────────────────
        if phase == 'STRAIGHT':
            delta = 0.0
            if odist >= straight_dist:
                phase = 'EVADE'
                phase_start = odist

        # EVADE: 회피 기동  ────────────────────────────────────
        elif phase == 'EVADE':
            delta = delta_evade
            if odist - phase_start >= EVADE_DIST:
                phase = 'PD_RETURN'
                phase_start    = odist
                pd_start_od    = odist
                # PD 초기 오차 저장 (D항 스파이크 방지 — JS 동일)
                prev_pd_err    = (theta - TARGET_THETA) * (180 / math.pi)

        # PD_RETURN: 헤딩 복귀  ────────────────────────────────
        #  JS pdControl(): steer_py = clip(-err·Kp_eff − rate·Kd_eff)
        #                  steer_gui = −steer_py   (부호 반전: Python→GUI)
        #  updateWheelDeg(steer_gui): raw = −steer_gui·(π/180)
        elif phase == 'PD_RETURN':
            err_deg  = (theta - TARGET_THETA) * (180 / math.pi)
            err_rate = (err_deg - prev_pd_err) / PHYS_DT
            scale    = PD_VREF / max(v, 1.0)
            steer_py = max(-30.0, min(30.0,
                           (-err_deg  * PD_KP * scale)
                           - (err_rate * PD_KD * scale)))
            steer_gui = -steer_py                    # Python→GUI 부호 변환
            prev_pd_err = err_deg

            raw   = -steer_gui * math.pi / 180.0    # updateWheelDeg 내부 로직
            delta = max(-MAX_STEER_RAD, min(MAX_STEER_RAD, raw))

            # 수렴 조건 — JS: abs(errDeg)<1.0 && abs(steerGUI)<0.5
            if abs(err_deg) < 1.0 and abs(steer_gui) < 0.5:
                phase = 'LANE_RETURN'
                phase_start    = odist
                lane_start_od  = odist
                # prevLaneErr 초기화 (JS 동일: D항 스파이크 방지)
                la_init  = max(10.0, v * 2.0)
                dt_init  = math.pi / 2 + math.atan2(x, la_init)
                prev_lane_err = (theta - dt_init) * (180 / math.pi)

        # LANE_RETURN: 차선 복귀 — Look-Ahead Heading Controller  ─
        #  JS: desiredTheta = π/2 + atan2(xErr, lookAhead)
        #      steerCmd = hdgErr·Kp·scale + errRate·Kd·scale
        #      updateWheelDeg(steerCmd)
        elif phase == 'LANE_RETURN':
            la           = max(10.0, v * 2.0)
            desired_th   = math.pi / 2 + math.atan2(x, la)
            hdg_err      = (theta - desired_th) * (180 / math.pi)
            scale        = PD_VREF / max(v, 1.0)
            err_rate     = (hdg_err - prev_lane_err) / PHYS_DT
            steer_gui    = max(-30.0, min(30.0,
                               hdg_err * PD_KP * scale
                               + err_rate * PD_KD * scale))
            prev_lane_err = hdg_err

            raw   = -steer_gui * math.pi / 180.0
            delta = max(-MAX_STEER_RAD, min(MAX_STEER_RAD, raw))

            # 수렴 조건 — JS: abs(x)<0.3 && abs(heading_err)<0.5
            # ★ passed_obs 조건 추가: 장애물 z-위치를 통과한 뒤에만 진정한 SUCCESS
            #   (직진 0° 등에서 장애물 도달 전에 일찍 수렴하는 허위 성공 방지)
            s_hdg = (theta - TARGET_THETA) * (180 / math.pi)
            if abs(x) < 0.3 and abs(s_hdg) < 0.5 and passed_obs:
                return _row(steer_sw, straight_dist, speed_kmh,
                            'SUCCESS', 1,
                            min_obs_dist - COLL_RADIUS, max_lat_dev,
                            lat_at_obs,
                            pd_start_od, lane_start_od, odist)

        # ── 기구학적 자전거 모델 업데이트 (JS 와 동일) ─────────────
        theta += (v / WHEELBASE) * math.tan(delta) * PHYS_DT
        x     += v * math.cos(theta) * PHYS_DT
        z     -= v * math.sin(theta) * PHYS_DT
        odist += v * PHYS_DT

        if abs(x) > max_lat_dev:
            max_lat_dev = abs(x)

    return _row(steer_sw, straight_dist, speed_kmh,
                'TIMEOUT', 0,
                min_obs_dist - COLL_RADIUS, max_lat_dev,
                lat_at_obs if passed_obs else x,
                pd_start_od, lane_start_od, odist)


def _row(steer_sw, straight_dist, speed_kmh, outcome, success,
         clearance, lat_dev, lat_obs, pd_od, lane_od, total_od):
    return {
        'speed_kmh':         speed_kmh,
        'steer_deg_sw':      steer_sw,
        'steer_deg_wheel':   round(steer_sw / STEERING_RATIO, 2),
        'straight_dist_m':   straight_dist,
        'avoidance_success': success,
        'outcome':           outcome,
        'min_clearance_m':   round(clearance, 3),
        'max_lat_dev_m':     round(lat_dev,   3),
        'lat_at_obstacle_m': round(lat_obs,   3),
        'pd_start_dist_m':   round(pd_od,   1) if pd_od   >= 0 else 'N/A',
        'lane_start_dist_m': round(lane_od, 1) if lane_od >= 0 else 'N/A',
        'total_dist_m':      round(total_od, 1),
    }


# ══════════════════════════════════════════════════════════════
#  병렬 스윕
# ══════════════════════════════════════════════════════════════
def run_sweep(params: Dict, n_workers: int = 0) -> List[Dict]:
    args = [(st, sd, sp)
            for sp in params['speed_kmh']
            for sd in params['straight_dist_m']
            for st in params['steer_deg']]
    total = len(args)

    if n_workers <= 0:
        n_workers = min(multiprocessing.cpu_count(), 8)

    print(f"  총 시뮬레이션: {total:,}  (워커 {n_workers}개)")
    t0    = time.time()
    rows  = []

    with multiprocessing.Pool(n_workers) as pool:
        for row in pool.imap_unordered(simulate, args, chunksize=200):
            rows.append(row)
            n = len(rows)
            if n % 2000 == 0:
                print(f"    {n:>6,}/{total:,}  {n/total*100:.1f}%")

    print(f"  완료: {len(rows):,}  ({time.time()-t0:.1f}s)")
    return rows


# ══════════════════════════════════════════════════════════════
#  집계 헬퍼
# ══════════════════════════════════════════════════════════════
def pivot(rows, row_key, col_key, value_fn=None):
    """rows를 row_key × col_key 피벗으로 집계. value_fn(group) → 값."""
    from collections import defaultdict
    groups: Dict[tuple, list] = defaultdict(list)
    for r in rows:
        groups[(r[row_key], r[col_key])].append(r)

    row_vals = sorted(set(r[row_key] for r in rows))
    col_vals = sorted(set(r[col_key] for r in rows))
    mat = np.full((len(row_vals), len(col_vals)), np.nan)

    for ri, rv in enumerate(row_vals):
        for ci, cv in enumerate(col_vals):
            g = groups.get((rv, cv), [])
            if g:
                mat[ri, ci] = (value_fn(g) if value_fn
                               else sum(r['avoidance_success'] for r in g) / len(g) * 100)
    return row_vals, col_vals, mat


# ══════════════════════════════════════════════════════════════
#  시각화
# ══════════════════════════════════════════════════════════════
SUCCESS_CMAP = LinearSegmentedColormap.from_list(
    'sr', ['#1a0a2e', '#1e3a6e', '#1a6ade', '#00c8a0', '#3dffaa'], N=256)

def _heatmap(ax, mat, row_vals, col_vals, title, xlabel, ylabel, fmt='.0f'):
    im = ax.imshow(mat, aspect='auto', origin='lower',
                   cmap=SUCCESS_CMAP, vmin=0, vmax=100,
                   extent=[col_vals[0], col_vals[-1],
                            row_vals[0], row_vals[-1]])
    ax.set_title(title, fontsize=11, fontweight='bold', pad=8)
    ax.set_xlabel(xlabel, fontsize=9)
    ax.set_ylabel(ylabel, fontsize=9)
    ax.tick_params(labelsize=8)
    cb = plt.colorbar(im, ax=ax, fraction=0.046, pad=0.04)
    cb.set_label('Success %', fontsize=8)
    cb.ax.tick_params(labelsize=7)


def make_plots(rows: List[Dict], out_path: str):
    speeds  = sorted(set(r['speed_kmh']       for r in rows))
    steers  = sorted(set(r['steer_deg_sw']     for r in rows))
    dists   = sorted(set(r['straight_dist_m']  for r in rows))

    # ── Figure 1: 히트맵 3 + 라인 3 ──────────────────────────
    fig = plt.figure(figsize=(18, 11))
    fig.patch.set_facecolor('#0d1220')
    gs  = gridspec.GridSpec(2, 3, figure=fig,
                            hspace=0.40, wspace=0.38,
                            top=0.91, bottom=0.07, left=0.06, right=0.97)

    ax_list = [fig.add_subplot(gs[i, j]) for i in range(2) for j in range(3)]
    for ax in ax_list:
        ax.set_facecolor('#111828')
        ax.tick_params(colors='#8aaccc', labelsize=8)
        ax.xaxis.label.set_color('#8aaccc')
        ax.yaxis.label.set_color('#8aaccc')
        ax.title.set_color('#c8e0ff')
        for sp in ax.spines.values():
            sp.set_edgecolor('#1e3a58')

    # ── (0,0) 히트맵: 조향각 × 기동거리  (속도 40 km/h 기준) ──
    v_ref = min(speeds, key=lambda s: abs(s - 40))
    sub   = [r for r in rows if r['speed_kmh'] == v_ref]
    rv, cv, mat = pivot(sub, 'straight_dist_m', 'steer_deg_sw')
    ax = ax_list[0]
    _heatmap(ax, mat, rv, cv,
             f'Steer x Evade-Start  (v={v_ref} km/h)',
             'Steer SW (deg)', 'Evade Start (m)')

    # ── (0,1) Heatmap: steer x speed  (start_dist=30 m) ──
    d_ref = min(dists, key=lambda d: abs(d - 30))
    sub   = [r for r in rows if r['straight_dist_m'] == d_ref]
    rv, cv, mat = pivot(sub, 'speed_kmh', 'steer_deg_sw')
    ax = ax_list[1]
    _heatmap(ax, mat, rv, cv,
             f'Steer x Speed  (start={d_ref} m)',
             'Steer SW (deg)', 'Speed (km/h)')

    # ── (0,2) Heatmap: start_dist x speed  (steer ~45 deg) ──
    s_ref = min([s for s in steers if s > 0], key=lambda s: abs(s - 45))
    sub   = [r for r in rows if r['steer_deg_sw'] == s_ref]
    rv, cv, mat = pivot(sub, 'speed_kmh', 'straight_dist_m')
    ax = ax_list[2]
    _heatmap(ax, mat, rv, cv,
             f'Start-Dist x Speed  (steer={s_ref} deg)',
             'Evade Start (m)', 'Speed (km/h)')

    # ── (1,0) Bar: success rate by speed ──
    ax = ax_list[3]
    spd_rates = []
    for sp in speeds:
        g = [r for r in rows if r['speed_kmh'] == sp]
        spd_rates.append(sum(r['avoidance_success'] for r in g) / len(g) * 100)

    bars = ax.bar(speeds, spd_rates, width=7, color='#2a7ade', alpha=0.85,
                  edgecolor='#1a5aae')
    for bar, val in zip(bars, spd_rates):
        ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 1,
                f'{val:.0f}%', ha='center', va='bottom',
                fontsize=7, color='#8aaccc')
    ax.set_title('Success Rate by Speed', fontsize=11, fontweight='bold')
    ax.set_xlabel('Speed (km/h)', fontsize=9)
    ax.set_ylabel('Success Rate (%)', fontsize=9)
    ax.set_ylim(0, 115)
    ax.axhline(50, color='#ff6644', lw=0.8, ls='--', alpha=0.5)

    # ── (1,1) Line: success rate vs steer angle, grouped by speed ──
    ax = ax_list[4]
    sel_speeds = speeds[::max(1, len(speeds)//6)][:6]
    palette = ['#3dffaa', '#00d4ff', '#2a7ade', '#ffaa00', '#ff6644', '#ff3366']
    for sp, col in zip(sel_speeds, palette):
        g   = [r for r in rows if r['speed_kmh'] == sp]
        pts = [(st, sum(r['avoidance_success'] for r in g
                        if r['steer_deg_sw'] == st)
                    / max(1, sum(1 for r in g if r['steer_deg_sw'] == st)) * 100)
               for st in steers]
        xs2, ys2 = zip(*pts)
        ax.plot(xs2, ys2, color=col, lw=1.5, label=f'{sp} km/h', alpha=0.9)

    ax.set_title('Success Rate vs Steer Angle', fontsize=11, fontweight='bold')
    ax.set_xlabel('Steer SW (deg)', fontsize=9)
    ax.set_ylabel('Success Rate (%)', fontsize=9)
    ax.legend(fontsize=7, loc='upper left',
              facecolor='#1a2a40', edgecolor='#1e3a58', labelcolor='#8aaccc')
    ax.axhline(50, color='#ffffff', lw=0.5, ls='--', alpha=0.3)
    ax.set_ylim(-5, 110)

    # ── (1,2) Line: success rate vs evade start dist, grouped by speed ──
    ax = ax_list[5]
    for sp, col in zip(sel_speeds, palette):
        g   = [r for r in rows if r['speed_kmh'] == sp]
        pts = [(d, sum(r['avoidance_success'] for r in g
                       if r['straight_dist_m'] == d)
                   / max(1, sum(1 for r in g if r['straight_dist_m'] == d)) * 100)
               for d in dists]
        xs2, ys2 = zip(*pts)
        ax.plot(xs2, ys2, color=col, lw=1.5, label=f'{sp} km/h', alpha=0.9)

    ax.axvline(abs(OBS_Z) - COLL_RADIUS, color='#ff4444', lw=1.0, ls=':',
               alpha=0.7, label=f'Obs boundary {abs(OBS_Z)-COLL_RADIUS:.0f}m')
    ax.set_title('Success Rate vs Evade Start', fontsize=11, fontweight='bold')
    ax.set_xlabel('Evade Start (m)', fontsize=9)
    ax.set_ylabel('Success Rate (%)', fontsize=9)
    ax.legend(fontsize=7, loc='upper right',
              facecolor='#1a2a40', edgecolor='#1e3a58', labelcolor='#8aaccc')
    ax.set_ylim(-5, 110)

    fig.suptitle(
        'Autonomous Driving Simulator — Avoidance Success Rate Analysis\n'
        f'Obstacle: x={OBS_X} z={OBS_Z} m  |  '
        f'Coll.radius={COLL_RADIUS} m  |  '
        f'Cliff |x|>{CLIFF_X} m  |  '
        f'EVADE={EVADE_DIST} m  |  SR={STEERING_RATIO}',
        fontsize=12, fontweight='bold', color='#c8e0ff', y=0.97)

    plt.savefig(out_path, dpi=150, bbox_inches='tight',
                facecolor='#0d1220')
    plt.close(fig)
    print(f"  그래프: {out_path}")


# ══════════════════════════════════════════════════════════════
#  Excel 출력
# ══════════════════════════════════════════════════════════════
_HDR_FILL  = PatternFill('solid', fgColor='0d2240')
_HDR_FONT  = Font(bold=True, color='6ec6ff', size=10)
_SUB_FILL  = PatternFill('solid', fgColor='080f20')
_ALT_FILL  = PatternFill('solid', fgColor='0a1428')
_OK_FILL   = PatternFill('solid', fgColor='0a2a14')
_NG_FILL   = PatternFill('solid', fgColor='2a0808')
_TH_BORDER = Border(
    bottom=Side(style='thin', color='1e4a7a'),
    top=Side(style='thin', color='1e4a7a'))

def _hdr(ws, row, cols):
    for c, val in enumerate(cols, 1):
        cell = ws.cell(row, c, val)
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = _TH_BORDER


def _cell(ws, row, col, val, bold=False, color='c8e0ff', fill=None, align='center'):
    c = ws.cell(row, col, val)
    c.font = Font(bold=bold, color=color, size=9)
    c.alignment = Alignment(horizontal=align, vertical='center')
    if fill:
        c.fill = fill
    return c


def _pct_color(pct: float) -> str:
    """성공률 → 셀 배경색 (0%=어두운 빨강, 100%=어두운 초록)"""
    r = int(10  + (1 - pct) * 50)
    g = int(10  +       pct * 50)
    b = 20
    return f'{r:02x}{g:02x}{b:02x}'


def save_excel(rows: List[Dict], out_path: str):
    wb = openpyxl.Workbook()

    speeds = sorted(set(r['speed_kmh']       for r in rows))
    steers = sorted(set(r['steer_deg_sw']     for r in rows))
    dists  = sorted(set(r['straight_dist_m']  for r in rows))

    # ── Sheet 1: 원시 데이터 ───────────────────────────────────
    ws = wb.active
    ws.title = '원시 데이터'
    ws.sheet_properties.tabColor = '1a6ade'
    ws.freeze_panes = 'A2'
    cols = list(rows[0].keys())
    _hdr(ws, 1, cols)
    for ri, r in enumerate(rows, 2):
        for ci, k in enumerate(cols, 1):
            val = r[k]
            fill = _OK_FILL if k == 'avoidance_success' and val == 1 else \
                   _NG_FILL if k == 'avoidance_success' and val == 0 else None
            _cell(ws, ri, ci, val,
                  fill=fill if fill else (_ALT_FILL if ri % 2 == 0 else None),
                  align='center')
    for ci, k in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = 16

    # ── Sheet 2: 속도별 성공률 ────────────────────────────────
    ws2 = wb.create_sheet('속도별 성공률')
    ws2.sheet_properties.tabColor = '3dffaa'
    headers = ['속도 (km/h)', '총 시뮬', '성공', '실패', '성공률 (%)',
               '충돌', '낭떠러지', '타임아웃']
    _hdr(ws2, 1, headers)
    for ri, sp in enumerate(speeds, 2):
        g = [r for r in rows if r['speed_kmh'] == sp]
        ok  = sum(r['avoidance_success'] for r in g)
        col = sum(1 for r in g if 'COLLISION' in r['outcome'])
        clf = sum(1 for r in g if 'CLIFF' in r['outcome'])
        tmo = sum(1 for r in g if r['outcome'] == 'TIMEOUT')
        pct = ok / len(g) * 100
        vals = [sp, len(g), ok, len(g)-ok, round(pct, 1), col, clf, tmo]
        fill = PatternFill('solid', fgColor=_pct_color(pct / 100))
        for ci, v in enumerate(vals, 1):
            _cell(ws2, ri, ci, v, fill=fill)
    for ci in range(1, 9):
        ws2.column_dimensions[get_column_letter(ci)].width = 15

    # ── Sheet 3: 조향각별 성공률 ──────────────────────────────
    ws3 = wb.create_sheet('조향각별 성공률')
    ws3.sheet_properties.tabColor = '00d4ff'
    _hdr(ws3, 1, ['핸들 조향각 SW (°)', '바퀴 조향각 (°)',
                  '총 시뮬', '성공', '성공률 (%)'])
    for ri, st in enumerate(steers, 2):
        g   = [r for r in rows if r['steer_deg_sw'] == st]
        ok  = sum(r['avoidance_success'] for r in g)
        pct = ok / len(g) * 100 if g else 0
        fill = PatternFill('solid', fgColor=_pct_color(pct / 100))
        vals = [st, round(st / STEERING_RATIO, 1), len(g), ok, round(pct, 1)]
        for ci, v in enumerate(vals, 1):
            _cell(ws3, ri, ci, v, fill=fill)
    for ci in range(1, 6):
        ws3.column_dimensions[get_column_letter(ci)].width = 18

    # ── Sheet 4: 기동거리별 성공률 ────────────────────────────
    ws4 = wb.create_sheet('기동거리별 성공률')
    ws4.sheet_properties.tabColor = 'ffaa00'
    _hdr(ws4, 1, ['기동시작점 (m)', '총 시뮬', '성공', '성공률 (%)',
                  '비고'])
    for ri, d in enumerate(dists, 2):
        g   = [r for r in rows if r['straight_dist_m'] == d]
        ok  = sum(r['avoidance_success'] for r in g)
        pct = ok / len(g) * 100 if g else 0
        fill = PatternFill('solid', fgColor=_pct_color(pct / 100))
        note = '장애물 도달 전 기동 불가' if d >= abs(OBS_Z) - COLL_RADIUS else ''
        vals = [d, len(g), ok, round(pct, 1), note]
        for ci, v in enumerate(vals, 1):
            _cell(ws4, ri, ci, v, fill=fill, align='center' if ci < 5 else 'left')
    for ci in range(1, 6):
        ws4.column_dimensions[get_column_letter(ci)].width = 18

    # ── Sheet 5: 피벗 — 조향각 × 속도 성공률 (%) ─────────────
    ws5 = wb.create_sheet('피벗_조향×속도')
    ws5.sheet_properties.tabColor = 'b060ff'
    ws5.cell(1, 1, '조향각(°) \\ 속도(km/h)')
    ws5.cell(1, 1).fill = _HDR_FILL
    ws5.cell(1, 1).font = Font(bold=True, color='b060ff', size=9)
    ws5.cell(1, 1).alignment = Alignment(horizontal='center')
    for ci, sp in enumerate(speeds, 2):
        c = ws5.cell(1, ci, sp)
        c.fill = _HDR_FILL; c.font = _HDR_FONT
        c.alignment = Alignment(horizontal='center')
    for ri, st in enumerate(steers, 2):
        c = ws5.cell(ri, 1, st)
        c.fill = _HDR_FILL; c.font = _HDR_FONT
        c.alignment = Alignment(horizontal='center')
        for ci, sp in enumerate(speeds, 2):
            g   = [r for r in rows if r['steer_deg_sw'] == st and r['speed_kmh'] == sp]
            pct = sum(r['avoidance_success'] for r in g) / len(g) * 100 if g else 0
            fill = PatternFill('solid', fgColor=_pct_color(pct / 100))
            c2  = ws5.cell(ri, ci, round(pct, 0))
            c2.fill = fill
            c2.font = Font(bold=pct >= 50, color='c8e0ff' if pct >= 30 else '553333',
                           size=9)
            c2.alignment = Alignment(horizontal='center')
    ws5.column_dimensions['A'].width = 20
    for ci in range(2, len(speeds) + 2):
        ws5.column_dimensions[get_column_letter(ci)].width = 10

    # ── Sheet 6: 피벗 — 기동거리 × 속도 성공률 (%) ───────────
    ws6 = wb.create_sheet('피벗_거리×속도')
    ws6.sheet_properties.tabColor = 'ff6644'
    ws6.cell(1, 1, '기동거리(m) \\ 속도(km/h)')
    ws6.cell(1, 1).fill = _HDR_FILL
    ws6.cell(1, 1).font = Font(bold=True, color='ff8866', size=9)
    ws6.cell(1, 1).alignment = Alignment(horizontal='center')
    for ci, sp in enumerate(speeds, 2):
        c = ws6.cell(1, ci, sp)
        c.fill = _HDR_FILL; c.font = _HDR_FONT
        c.alignment = Alignment(horizontal='center')
    for ri, d in enumerate(dists, 2):
        c = ws6.cell(ri, 1, d)
        c.fill = _HDR_FILL; c.font = Font(bold=True, color='ff8866', size=9)
        c.alignment = Alignment(horizontal='center')
        for ci, sp in enumerate(speeds, 2):
            g   = [r for r in rows if r['straight_dist_m'] == d and r['speed_kmh'] == sp]
            pct = sum(r['avoidance_success'] for r in g) / len(g) * 100 if g else 0
            fill = PatternFill('solid', fgColor=_pct_color(pct / 100))
            c2  = ws6.cell(ri, ci, round(pct, 0))
            c2.fill = fill
            c2.font = Font(bold=pct >= 50, color='c8e0ff' if pct >= 30 else '553333',
                           size=9)
            c2.alignment = Alignment(horizontal='center')
    ws6.column_dimensions['A'].width = 20
    for ci in range(2, len(speeds) + 2):
        ws6.column_dimensions[get_column_letter(ci)].width = 10

    wb.save(out_path)
    print(f"  엑셀:  {out_path}  ({len(rows):,}행, {len(wb.sheetnames)}시트)")


# ══════════════════════════════════════════════════════════════
#  텍스트 리포트
# ══════════════════════════════════════════════════════════════
def build_report(rows: List[Dict]) -> str:
    from collections import Counter
    L = []
    sep = '═' * 66

    total = len(rows)
    ok    = sum(r['avoidance_success'] for r in rows)
    clf   = sum(1 for r in rows if 'CLIFF' in r['outcome'])
    coll  = sum(1 for r in rows if 'COLLISION' in r['outcome'])
    tmo   = sum(1 for r in rows if r['outcome'] == 'TIMEOUT')

    speeds = sorted(set(r['speed_kmh']      for r in rows))
    steers = sorted(set(r['steer_deg_sw']    for r in rows))
    dists  = sorted(set(r['straight_dist_m'] for r in rows))

    L += [sep,
          '  자율주행 시뮬레이터  회피 기동 분석 리포트  v2.0',
          sep,
          '',
          '  ▶ 물리 상수 (index.html 완전 동일)',
          f'    축거 L            : {WHEELBASE} m',
          f'    최대 바퀴 조향각  : ±{math.degrees(MAX_STEER_RAD):.0f}°'
          f'  (핸들 ±{math.degrees(MAX_STEER_RAD)*STEERING_RATIO:.0f}°)',
          f'    조향 비율 SR      : {STEERING_RATIO}',
          f'    충돌 반경         : {COLL_RADIUS} m'
          f' (차량 {VEH_RADIUS} + 장애물 {OBS_RADIUS})',
          f'    낭떠러지 경계     : |x| > {CLIFF_X} m',
          f'    EVADE 구간        : {EVADE_DIST} m',
          f'    물리 DT           : {PHYS_DT} s  (50 Hz)',
          f'    장애물 위치       : x={OBS_X}, z={OBS_Z} m  (JS DEFAULT_OBS)',
          f'    PD Kp / Kd        : {PD_KP} / {PD_KD}'
          f'  (VREF={PD_VREF*3.6:.0f} km/h)',
          '',
          '  ▶ 파라미터 공간 (= index.html GUI 슬라이더 3개)',
          f'    핸들 조향각 범위  : {min(steers)} ~ {max(steers)} ° SW'
          f'  ({len(steers)} 단계)',
          f'    기동 시작점 범위  : {min(dists)} ~ {max(dists)} m'
          f'  ({len(dists)} 단계)',
          f'    속도 범위         : {min(speeds)} ~ {max(speeds)} km/h'
          f'  ({len(speeds)} 단계)',
          f'    총 시뮬레이션     : {total:,} 회',
          '',
          '  ▶ 전체 결과',
          f'    SUCCESS           : {ok:>8,}  ({ok/total*100:.1f}%)',
          f'    COLLISION         : {coll:>8,}  ({coll/total*100:.1f}%)',
          f'    CLIFF             : {clf:>8,}  ({clf/total*100:.1f}%)',
          f'    TIMEOUT           : {tmo:>8,}  ({tmo/total*100:.1f}%)',
          '']

    # 속도별 성공률
    L += ['  ▶ 속도별 성공률',
          f"    {'속도':>8}  {'총수':>7}  {'성공':>7}  {'성공률':>7}"
          f"  {'충돌':>6}  {'낭떠러지':>8}",
          '    ' + '─' * 54]
    for sp in speeds:
        g   = [r for r in rows if r['speed_kmh'] == sp]
        sok = sum(r['avoidance_success'] for r in g)
        sc  = sum(1 for r in g if 'COLLISION' in r['outcome'])
        sf  = sum(1 for r in g if 'CLIFF' in r['outcome'])
        L.append(f'    {sp:>5} km/h  {len(g):>7,}  {sok:>7,}'
                 f'  {sok/len(g)*100:>6.1f}%  {sc:>6,}  {sf:>8,}')

    # 조향각별 성공률 (상위/하위)
    steer_rates = []
    for st in steers:
        g = [r for r in rows if r['steer_deg_sw'] == st]
        if g:
            steer_rates.append((st, sum(r['avoidance_success'] for r in g)
                                     / len(g) * 100))
    best3  = sorted(steer_rates, key=lambda x: -x[1])[:3]
    worst3 = sorted(steer_rates, key=lambda x:  x[1])[:3]
    L += ['',
          '  ▶ 최고 조향각 (전체 평균 성공률 기준)',
          *(f'    {st:>+4}° SW ({st/STEERING_RATIO:+.1f}° wheel)  → {r:.1f}%'
            for st, r in best3),
          '',
          '  ▶ 최저 조향각',
          *(f'    {st:>+4}° SW ({st/STEERING_RATIO:+.1f}° wheel)  → {r:.1f}%'
            for st, r in worst3)]

    L += ['', sep]
    return '\n'.join(L)


# ══════════════════════════════════════════════════════════════
#  CSV 저장
# ══════════════════════════════════════════════════════════════
def save_csv(rows: List[Dict], path: str):
    if not rows:
        return
    with open(path, 'w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=rows[0].keys())
        w.writeheader()
        w.writerows(rows)
    print(f"  CSV:   {path}  ({len(rows):,}행)")


# ══════════════════════════════════════════════════════════════
#  메인
# ══════════════════════════════════════════════════════════════
def main():
    if '--help' in sys.argv or '-h' in sys.argv:
        print(__doc__)
        sys.exit(0)

    quick    = '--quick' in sys.argv
    params   = PARAM_QUICK if quick else PARAM_FULL

    n_workers = 0
    if '--workers' in sys.argv:
        idx = sys.argv.index('--workers')
        try:
            n_workers = int(sys.argv[idx + 1])
        except (IndexError, ValueError):
            pass

    mode = 'QUICK' if quick else 'FULL'
    print()
    print('━' * 55)
    print('  자율주행 시뮬레이터  회피 기동 데이터셋  v2.0')
    print(f'  모드: {mode}   장애물: x={OBS_X}, z={OBS_Z} m')
    print('━' * 55)

    print('\n[1/4] 파라미터 스윕 실행 중...')
    rows = run_sweep(params, n_workers)

    base = os.path.dirname(os.path.abspath(__file__))

    print('\n[2/4] CSV 저장...')
    save_csv(rows, os.path.join(base, 'avoidance_dataset.csv'))

    print('\n[3/4] 엑셀 저장...')
    save_excel(rows, os.path.join(base, 'avoidance_dataset.xlsx'))

    print('\n[4/4] 그래프 + 리포트 저장...')
    make_plots(rows, os.path.join(base, 'avoidance_plots.png'))

    report = build_report(rows)
    rpath  = os.path.join(base, 'avoidance_report.txt')
    with open(rpath, 'w', encoding='utf-8') as f:
        f.write(report)
    print(f'  리포트: {rpath}')

    print()
    print(report)


if __name__ == '__main__':
    multiprocessing.freeze_support()
    main()
